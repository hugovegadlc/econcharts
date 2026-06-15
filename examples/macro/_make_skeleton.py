"""Generate macro.xlsx skeleton for peru_macro.yaml.

Formula hints go on header cell comments (Excel hover tooltips), NOT as a
data row — pd.read_excel(header=0) would treat row 2 as data, causing
_coerce_periods to choke on the empty-string period value.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from datetime import date
from pathlib import Path

wb = Workbook()
del wb[wb.sheetnames[0]]   # remove the default Sheet

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="0000A3")   # macro blue
COMMENT_AUTHOR = "econcharts"


def make_sheet(name, period_col, periods, data_cols):
    ws = wb.create_sheet(name)
    all_cols = [period_col] + list(data_cols)
    hints = ["period column (do not edit)"] + list(data_cols.values())
    # row 1 — headers with formula hints as cell comments
    for c, (col, hint) in enumerate(zip(all_cols, hints), 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.comment   = Comment(hint, COMMENT_AUTHOR)
    # data rows — period column only, data columns left empty
    for r, period in enumerate(periods, 2):
        cell = ws.cell(row=r, column=1, value=period)
        if isinstance(period, date):
            cell.number_format = "YYYY-MM-DD"
    # column widths
    ws.column_dimensions["A"].width = 14
    for c in range(2, len(all_cols) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    return ws


# ── anual (1900-2025) ──────────────────────────────────────────────────────
make_sheet("anual", "year", range(1900, 2026), {
    # chart 1 — levels
    "pbi_bn":        "pbi / 1000  (miles de millones S/ 2007)",
    "cons_priv_bn":  "cons_priv / 1000",
    "inv_priv_bn":   "inv_bruta_fija_priv / 1000",
    # chart 2 — logarithms
    "log_pbi":       "LN(pbi)",
    "log_cons_priv": "LN(cons_priv)",
    "log_inv_priv":  "LN(inv_bruta_fija_priv)",
    # charts 3/7/9 — growth rates
    "pbi_g":         "100*(pbi/pbi_prev - 1)",
    "cons_priv_g":   "100*(cons_priv/cons_priv_prev - 1)",
    "inv_priv_g":    "100*(inv_bruta_fija_priv/inv_prev - 1)",
    "ti_g":          "100*(ti/ti_prev - 1)  [terms of trade]",
    # chart 4 — % GDP
    "cons_priv_pct": "100 * cons_priv / pbi",
    "inv_priv_pct":  "100 * inv_bruta_fija_priv / pbi",
    "gasto_pub_pct": "100 * (cons_pub + inv_bruta_fija_pub) / pbi",
    "balanza_c_pct": "100 * (export - import) / pbi",
    # charts 5/6 — as-is from BCRP
    "inflacion":      "as-is from BCRP  (% anual)",
    # chart 6 — split: two blue line segments with a blank gap during hyperinflation
    #   inflacion_pre  blanks at the END once inflation exceeds 50% (no interior NaN)
    #   inflacion_post blanks at the START before 1993 (no interior NaN)
    #   → PCHIP never bridges the gap; the period 1975-1993 is naturally invisible
    "inflacion_pre":  "=IF(inflacion<=50, inflacion, \"\")  early history; blank once inflation > 50%",
    "inflacion_post": "=IF(year>=1993, inflacion, \"\")     modern era (1993 onwards)",
    # chart 8 — as-is from BCRP
    "def_fiscal":    "as-is from BCRP  (% PBI; deficit = positive value)",
    # chart C — pending
    "rin":           "RIN miles de millones USD  [PENDIENTE — BCRP estadisticas anuales]",
    # chart B — pending / commented-out
    "brecha_pbi":    "100*(LN(pbi) - HP_trend)  lambda=100  [PENDIENTE]",
})

# ── mensual (Dec-1993 to Mar-2026) ────────────────────────────────────────
months = [
    date(y, m, 1)
    for y in range(1993, 2027)
    for m in range(1, 13)
    if not (y == 1993 and m < 12) and not (y == 2026 and m > 3)
]
make_sheet("mensual", "month", months, {
    "tcn":     "as-is from BCRP  (soles por USD)",
    "tcn_var": "100*(tcn / tcn_mes_anterior - 1)  depreciacion mensual %",
})

# ── bdp (1950-2025) ───────────────────────────────────────────────────────
make_sheet("bdp", "year", range(1950, 2026), {
    "bal_com_pct":  "100 * bal_com  / pbi_nominal   (balanza comercial % PBI)",
    "bal_ser_pct":  "100 * bal_ser  / pbi_nominal   (servicios % PBI)",
    "ing_prim_pct": "100 * ing_prim / pbi_nominal   (renta de factores % PBI)",
    "ing_sec_pct":  "100 * ing_sec  / pbi_nominal   (transferencias % PBI)",
})

# ── fiscal (1970-2025) ────────────────────────────────────────────────────
make_sheet("fiscal", "year", range(1970, 2026), {
    "ingreso":    "ingresos del gobierno % PBI  (as-is from BCRP)",
    "gasto_corr": "gasto corriente % PBI  (as-is from BCRP)",
    "gasto_cap":  "gasto de capital % PBI  (as-is from BCRP)",
    "inter":      "intereses % PBI  (as-is from BCRP)",
    "otros":      "otros gastos % PBI  (as-is from BCRP)",
    "deuda_pub":  "deuda publica % PBI  [PENDIENTE — BCRP estadisticas anuales]",
})

# ── pbi_pc (1960-2024, pending chart A) ──────────────────────────────────
make_sheet("pbi_pc", "year", range(1960, 2025), {
    "peru":     "LN(PBI per capita USD constante)  WDI: NY.GDP.PCAP.KD  [PENDIENTE]",
    "chile":    "LN(PBI per capita USD constante)  [PENDIENTE]",
    "colombia": "LN(PBI per capita USD constante)  [PENDIENTE]",
    "mexico":   "LN(PBI per capita USD constante)  [PENDIENTE]",
})

# ── NOTAS ─────────────────────────────────────────────────────────────────
ws_n = wb.create_sheet("NOTAS")
ws_n.sheet_properties.tabColor = "F07000"
ws_n.column_dimensions["A"].width = 100
lines = [
    ("MACRO.XLSX — skeleton for examples/macro/peru_macro.yaml", True),
    ("", False),
    ("FLUJO DE TRABAJO", True),
    ("  1. Pega los datos del addin BCRP en cada hoja (la columna de periodo ya esta pre-llenada).", False),
    ("  2. Las columnas marcadas 'as-is' se usan directamente sin transformacion.", False),
    ("  3. Las columnas con formula se calculan en Excel antes de correr el batch.", False),
    ("     Pasa el cursor sobre cada encabezado de columna para ver la formula.", False),
    ("  4. Corre:  econcharts build peru_macro.yaml", False),
    ("", False),
    ("HOJAS Y RANGOS DE PERIODO", True),
    ("  anual   — anual 1900-2025   (columna 'year', entero)", False),
    ("  mensual — mensual dic-1993 a mar-2026  (columna 'month', fecha YYYY-MM-DD)", False),
    ("  bdp     — anual 1950-2025   (balanza de pagos corriente)", False),
    ("  fiscal  — anual 1970-2025   (datos fiscales)", False),
    ("  pbi_pc  — anual 1960-2024   [pendiente datos externos — chart A]", False),
    ("", False),
    ("COLUMNAS PENDIENTES (datos no disponibles en el Excel actual)", True),
    ("  rin       (anual)   — BCRP estadisticas anuales / reservas internacionales", False),
    ("  deuda_pub (fiscal)  — BCRP estadisticas anuales / sector publico", False),
    ("  pbi_pc/* (pbi_pc)   — WDI indicador NY.GDP.PCAP.KD  o  Maddison Project", False),
    ("  brecha_pbi (anual)  — requiere filtro HP (lambda=100) sobre LN(pbi)", False),
    ("", False),
    ("DIFERENCIAS VS EL SCRIPT R", True),
    ("  Graph 10 (TCN): el grafico de 2 paneles (facet_wrap) se divide en", False),
    ("    tcn_nivel + tcn_dep — econcharts no tiene facets aun.", False),
    ("  Graph 6 (inflacion recortada): infla_reciente usa dos series azules (inflacion_pre +", False),
    ("    inflacion_post) para dejar un hueco visible durante la hiperinflacion.", False),
]
for r, (text, bold) in enumerate(lines, 1):
    cell = ws_n.cell(row=r, column=1, value=text)
    cell.font = Font(bold=bold, size=10 if bold else 9)

out = Path(__file__).parent / "macro.xlsx"
wb.save(out)
print(f"OK  {out}")
